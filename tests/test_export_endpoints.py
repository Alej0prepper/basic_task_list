# test_export_endpoints.py
import csv
import io

def _seed(client):
    client.post("/tasks", json={"text": "deploy backend", "status": "DONE"})
    client.post("/tasks", json={"text": "revisar docs", "status": "pending"})
    client.post("/tasks", json={"text": "deploy frontend", "status": "DONE"})

def test_export_csv_basic(client):
    _seed(client)
    r = client.get("/tasks-export.csv?q=deploy&sort=done&dir=desc")
    assert r.status_code == 200, r.text
    assert "text/csv" in (r.headers.get("content-type") or "").lower()
    content = r.text
    # Cabecera y al menos una fila
    assert "ID,Text,Status,Tags,Created At" in content.splitlines()[0]
    rows = list(csv.reader(io.StringIO(content)))
    assert len(rows) >= 2

def test_export_xlsx_basic(client):
    _seed(client)
    r = client.get("/tasks-export.xlsx?q=deploy&sort=done&dir=desc")
    assert r.status_code == 200, r.text
    ctype = (r.headers.get("content-type") or "").lower()
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in ctype
    # XLSX es un ZIP, empieza con 'PK'
    assert r.content[:2] == b"PK"
    # Intentar abrir con openpyxl si está instalado
    try:
        import openpyxl  # type: ignore
        from io import BytesIO
        wb = openpyxl.load_workbook(filename=BytesIO(r.content))
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))[0:5]]
        assert headers[:5] == ["ID", "Text", "Status", "Tags", "Created At"]
        assert ws.max_row >= 2
    except ImportError:
        # si no está instalado, damos por bueno con la firma 'PK'
        pass

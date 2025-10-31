from typing import Optional
from datetime import timedelta, datetime
from io import BytesIO, StringIO
import os
import csv

from fastapi import FastAPI, Depends, HTTPException, Query, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import JSONResponse, RedirectResponse, HTMLResponse, StreamingResponse, PlainTextResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware

from sqlalchemy.orm import Session

from sqladmin import Admin, ModelView

from .settings import settings
from . import deps, schemas, models, crud, utils
from .database import engine, SessionLocal
from .admin_auth import AdminAuth

# -----------------------------------------------------------------------------
# App & CORS
# -----------------------------------------------------------------------------
app = FastAPI(title="Tasklist")

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.CORS_ORIGINS or ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Usa misma secret para cookie de sesión y para el backend del admin
ADMIN_SESSION_SECRET = os.getenv("ADMIN_SESSION_SECRET", settings.SECRET_KEY)
app.add_middleware(SessionMiddleware, secret_key=ADMIN_SESSION_SECRET, same_site="lax")

app.mount("/static", StaticFiles(directory="tasklist_app/static"), name="static")

# -----------------------------------------------------------------------------
# Admin UI (/admin) con sqladmin + autenticación
# -----------------------------------------------------------------------------
authentication_backend = AdminAuth(secret_key=ADMIN_SESSION_SECRET)
admin = Admin(app, engine, authentication_backend=authentication_backend)

class UserAdmin(ModelView, model=models.User):
    column_list = [
        models.User.id,
        models.User.email,
        models.User.created_at,
        models.User.updated_at,
    ]
    column_searchable_list = [models.User.email]
    column_sortable_list = [models.User.id, models.User.created_at, models.User.updated_at]
    name = "Usuario"
    name_plural = "Usuarios"
    icon = "fa-solid fa-user"

class TaskAdmin(ModelView, model=models.Task):
    column_list = [
        models.Task.id,
        models.Task.text,
        models.Task.status,
        models.Task.tags,
        models.Task.owner_id,
        models.Task.created_at,
        models.Task.updated_at,
    ]
    column_searchable_list = [models.Task.text, models.Task.status]
    column_sortable_list = [models.Task.id, models.Task.created_at, models.Task.updated_at]
    name = "Tarea"
    name_plural = "Tareas"
    icon = "fa-solid fa-list-check"

admin.add_view(UserAdmin)
admin.add_view(TaskAdmin)

# Plantillas (carpeta interna)
templates = Jinja2Templates(directory="tasklist_app/templates")

# -----------------------------------------------------------------------------
# Helpers de cookies (JWT)
# -----------------------------------------------------------------------------
COOKIE_NAME = "access_token"

def set_auth_cookie(response: RedirectResponse | JSONResponse, token: str):
    response.set_cookie(
        key=COOKIE_NAME,
        value=f"Bearer {token}",
        httponly=True,
        secure=False,
        samesite="lax",
        max_age=int(timedelta(minutes=settings.ACCESS_TOKEN_EXPIRE_MINUTES).total_seconds()),
        path="/",
    )

def clear_auth_cookie(response: RedirectResponse | JSONResponse):
    response.delete_cookie(COOKIE_NAME, path="/")

def get_token_from_cookie(request: Request) -> Optional[str]:
    raw = request.cookies.get(COOKIE_NAME)
    if not raw:
        return None
    if raw.startswith("Bearer "):
        return raw.split(" ", 1)[1]
    return raw

def current_user_from_cookie(request: Request, db: Session) -> Optional[models.User]:
    token = get_token_from_cookie(request)
    if not token:
        return None
    try:
        payload = utils.jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])  # type: ignore
        email = payload.get("sub")
        if not email:
            return None
        return db.query(models.User).filter(models.User.email == email).first()
    except Exception:
        return None

# -----------------------------------------------------------------------------
# Rutas de AUTH (API)
# -----------------------------------------------------------------------------
@app.post("/auth/register", response_model=schemas.UserOut, status_code=201)
def register(user_in: schemas.UserCreate, db: Session = Depends(deps.get_db)):
    exists = db.query(models.User).filter(models.User.email == user_in.email).first()
    if exists:
        raise HTTPException(status_code=400, detail="Email already registered")
    user = models.User(email=user_in.email, password_hash=utils.hash_password(user_in.password))
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

@app.post("/auth/login", response_model=schemas.Token)
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(deps.get_db)):
    user = db.query(models.User).filter(models.User.email == form.username).first()
    if not user or not utils.verify_password(form.password, user.password_hash):
        raise HTTPException(status_code=400, detail="Incorrect username or password")
    token = utils.create_access_token({"sub": user.email})
    return {"access_token": token, "token_type": "bearer"}

# -----------------------------------------------------------------------------
# Rutas de SALUD
# -----------------------------------------------------------------------------
@app.get("/health")
def health():
    return {"status": "ok"}

# -----------------------------------------------------------------------------
# Rutas de TASKS (API)
# -----------------------------------------------------------------------------
@app.post("/tasks", response_model=schemas.TaskOut, status_code=201)
def create_task(
    task_in: schemas.TaskCreate,
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
):
    return crud.create_task(db=db, task_in=task_in, owner_id=current_user.id)

@app.get("/tasks/{task_id}", response_model=schemas.TaskOut)
def get_task(task_id: int, db: Session = Depends(deps.get_db)):
    t = crud.get_task(db, task_id)
    if not t:
        raise HTTPException(status_code=404, detail="Not found")
    return t

@app.put("/tasks/{task_id}", response_model=schemas.TaskOut)
def update_task(task_id: int, task_in: schemas.TaskUpdate, db: Session = Depends(deps.get_db)):
    t = crud.update_task(db, task_id, task_in)
    if not t:
        raise HTTPException(status_code=404, detail="Not found")
    return t

@app.delete("/tasks/{task_id}", status_code=204)
def delete_task(task_id: int, db: Session = Depends(deps.get_db)):
    ok = crud.delete_task(db, task_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Not found")
    return {"detail": "deleted"}

@app.get("/tasks", response_model=schemas.PageTasks)
def list_tasks(
    limit: int = Query(10, ge=1, le=100),
    offset: int = Query(0, ge=0),
    status: Optional[str] = None,
    q: Optional[str] = Query(None, description="free-text over task text"),
    sort: Optional[str] = Query("date", description="date | done"),
    dir: Optional[str] = Query("desc", description="asc | desc"),
    db: Session = Depends(deps.get_db),
    current_user: Optional[models.User] = Depends(deps.get_current_user_optional),
):
    owner_id = current_user.id if current_user else None
    order_by = "done" if (sort or "").lower() == "done" else "created_at"
    order_dir = (dir or "desc")
    return crud.list_tasks_page(
        db=db,
        owner_id=owner_id,
        limit=limit,
        offset=offset,
        status=status,
        order_by=order_by,
        order_dir=order_dir,
        search=q,
    )

@app.get("/tasks-ui", response_model=schemas.PageTasks)
def list_tasks_ui(
    limit: int = Query(10, ge=1, le=100),
    offset: int = Query(0, ge=0),
    status: Optional[str] = None,
    q: Optional[str] = Query(None),
    sort: Optional[str] = Query("date"),
    dir: Optional[str] = Query("desc"),
    db: Session = Depends(deps.get_db),
    current_user: Optional[models.User] = Depends(deps.get_current_user_optional),
):
    owner_id = current_user.id if current_user else None
    order_by = "done" if (sort or "").lower() == "done" else "created_at"
    order_dir = (dir or "desc")
    return crud.list_tasks_page(
        db=db,
        owner_id=owner_id,
        limit=limit,
        offset=offset,
        status=status,
        order_by=order_by,
        order_dir=order_dir,
        search=q,
    )

# -----------------------------------------------------------------------------
# Rutas HTML (vista)
# -----------------------------------------------------------------------------
@app.get("/", include_in_schema=False)
def root_redirect():
    return RedirectResponse(url="/app")

@app.get("/app", response_class=HTMLResponse, include_in_schema=False)
def app_home(request: Request, db: Session = Depends(deps.get_db)):
    user = current_user_from_cookie(request, db)
    if user:
        return RedirectResponse(url="/app/tasks")
    return RedirectResponse(url="/app/login")

@app.get("/app/login", response_class=HTMLResponse, include_in_schema=False)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/app/login", response_class=HTMLResponse, include_in_schema=False)
def login_submit(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(deps.get_db),
):
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user or not utils.verify_password(password, user.password_hash):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Usuario o contraseña incorrectos."},
            status_code=400,
        )
    token = utils.create_access_token({"sub": user.email})
    resp = RedirectResponse(url="/app/tasks", status_code=302)
    set_auth_cookie(resp, token)
    return resp

@app.get("/app/logout", include_in_schema=False)
def logout():
    resp = RedirectResponse(url="/app/login", status_code=302)
    clear_auth_cookie(resp)
    return resp

@app.get("/app/tasks", response_class=HTMLResponse, include_in_schema=False)
def tasks_page(request: Request, db: Session = Depends(deps.get_db)):
    user = current_user_from_cookie(request, db)
    if not user:
        resp = RedirectResponse(url="/app/login", status_code=302)
        clear_auth_cookie(resp)
        return resp
    return templates.TemplateResponse("tasks.html", {"request": request, "user_email": user.email})

@app.get("/app/register", response_class=HTMLResponse, include_in_schema=False)
def register_page(request: Request):
    return templates.TemplateResponse("register.html", {"request": request})

@app.post("/app/register", response_class=HTMLResponse, include_in_schema=False)
def register_submit(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    password_confirm: str = Form(...),
    db: Session = Depends(deps.get_db),
):
    email_norm = (email or "").strip().lower()
    if not email_norm:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "El email es obligatorio.", "email_prefill": email},
            status_code=400,
        )
    if len(password or "") < 8:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "La contraseña debe tener al menos 8 caracteres.", "email_prefill": email},
            status_code=400,
        )
    if password != password_confirm:
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Las contraseñas no coinciden.", "email_prefill": email},
            status_code=400,
        )

    if crud.user_exists(db, email_norm):
        return templates.TemplateResponse(
            "register.html",
            {"request": request, "error": "Ese email ya está registrado.", "email_prefill": email},
            status_code=400,
        )

    user = crud.create_user(db, schemas.UserCreate(email=email_norm, password=password))
    token = utils.create_access_token({"sub": user.email})
    resp = RedirectResponse(url="/app/tasks", status_code=302)
    set_auth_cookie(resp, token)
    return resp

# -----------------------------------------------------------------------------
# Export: Excel y CSV
# -----------------------------------------------------------------------------
@app.get("/tasks-export.xlsx", include_in_schema=True)
def export_tasks_xlsx(
    status: Optional[str] = None,
    q: Optional[str] = Query(None, description="free-text search"),
    sort: Optional[str] = Query("date", description="date | done"),
    dir: Optional[str] = Query("desc", description="asc | desc"),
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
):
    owner_id = current_user.id if current_user else None
    order_by = "done" if (sort or "").lower() == "done" else "created_at"
    order_dir = (dir or "desc")

    items = crud.list_tasks_for_export(
        db=db,
        owner_id=owner_id,
        status=status,
        order_by=order_by,
        order_dir=order_dir,
        search=q,
    )

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = ["ID", "Text", "Status", "Tags", "Created At"]
    ws.append(headers)

    for t in items:
        tags = t.tags
        if isinstance(tags, (list, tuple)):
            tags_str = ", ".join(map(str, tags))
        else:
            tags_str = str(tags or "")
        ws.append([t.id, t.text, t.status, tags_str, t.created_at.isoformat() if t.created_at else ""])

    widths = [6, 60, 14, 30, 22]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    filename = f"tasks-{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/tasks-export.csv", response_class=PlainTextResponse, include_in_schema=True)
def export_tasks_csv(
    status: Optional[str] = None,
    q: Optional[str] = Query(None),
    sort: Optional[str] = Query("date"),
    dir: Optional[str] = Query("desc"),
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
):
    owner_id = current_user.id if current_user else None
    order_by = "done" if (sort or "").lower() == "done" else "created_at"
    order_dir = (dir or "desc")

    items = crud.list_tasks_for_export(
        db=db,
        owner_id=owner_id,
        status=status,
        order_by=order_by,
        order_dir=order_dir,
        search=q,
    )

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Text", "Status", "Tags", "Created At"])
    for t in items:
        tags = t.tags
        if isinstance(tags, (list, tuple)):
            tags_str = ", ".join(map(str, tags))
        else:
            tags_str = str(tags or "")
        writer.writerow([t.id, t.text, t.status, tags_str, t.created_at.isoformat() if t.created_at else ""])

    headers = {
        "Content-Disposition": f'attachment; filename="tasks-{datetime.utcnow().strftime("%Y%m%d-%H%M%S")}.csv"'
    }
    return PlainTextResponse(content=output.getvalue(), headers=headers, media_type="text/csv; charset=utf-8")

# # -----------------------------------------------------------------------------
# # Redirects legacy
# # -----------------------------------------------------------------------------
# @app.get("/tasks/ui", include_in_schema=False)
# def legacy_tasks_ui_redirect():
#     return RedirectResponse(url="/tasks-ui", status_code=307)
